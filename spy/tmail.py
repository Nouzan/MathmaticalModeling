import urllib.request
import http.cookiejar
import os
import pprint
import json
import re
import openpyxl
import shelve
import time
from bs4 import BeautifulSoup


def getHtml(url, refer):
    filename = 'cookie.txt'

    cj = http.cookiejar.MozillaCookieJar(filename)

    if os.path.exists(filename):
        cj.load(filename, ignore_discard=True, ignore_expires=True)

    if os.path.exists('subcookie.txt'):
        cookie = 'cna=CKu1EEICd14CATr5cCeTQeeq; l=AnNzLX2naYUMOVf9r0Vm2Uu/A32dsQcl; otherx=e%3D1%26p%3D*%26s%3D0%26c%3D0%26f%3D0%26g%3D0%26t%3D0; isg=AujoR3bpJZAsjgggVFXtxh2AuN-ZIUwbwsaamqIZLWMZ_YhnSiEcq34_gy7z; um=A502B1276E6D5FEF668B658F19DBB91D451BCFF1FA32C04E82530A2C4F4E1835EECF8A815DE1B301CD43AD3E795C914C3A8F490CFC9350A064C5556946B6DA5C; hng=CN%7Czh-CN%7CCNY; sm4=440303; _m_h5_tk=892a23775ff38263bef65048d6d7c623_1495005684418; _m_h5_tk_enc=991f2fd529aa8e89b35a5117ba2a61f4; t=40938fde33c1ea46c046d2a39ed59537; tracknick=nouzan%5Cu5C11%5Cu5E74; cookie2=1465a05d64d32fe3cc11c7a93f519cba; _tb_token_=fb816e437536; JSESSIONID=81FB4F01B8BD055B8C22FDCC0B01E811; uc1=cookie14=UoW%2BvfoThuLAJA%3D%3D&lng=zh_CN&cookie16=UIHiLt3xCS3yM2h4eKHS9lpEOw%3D%3D&existShop=false&cookie21=VFC%2FuZ9aiKIc&tag=8&cookie15=V32FPkk%2Fw0dUvg%3D%3D&pas=0; uc3=sg2=BxpQQly0fCG1cmljsFYTTGXK4E48CTK7jeFRhs5P4LQ%3D&nk2=DegrzmP4Ls8Uiw%3D%3D&id2=UUGrdwVcibd9yQ%3D%3D&vt3=F8dARVDXWenubKEURf4%3D&lg2=UtASsssmOIJ0bQ%3D%3D; uss=BdGo%2BUNTfrch8t%2BoBCfz5m%2BO28vUv%2BK3I3mWhRUjfljjA3Pyf1%2FttikN3Q%3D%3D; lgc=nouzan%5Cu5C11%5Cu5E74; sg=%E5%B9%B439; cookie1=BxubjUuEwYFMzp1VvQwUyUHgU9dGgEmxSZhM6dte4tQ%3D; unb=2991503703; _l_g_=Ug%3D%3D; _nk_=nouzan%5Cu5C11%5Cu5E74; cookie17=UUGrdwVcibd9yQ%3D%3D; login=true'
    else:
        cookie = 'ddd'

    proxy_support = urllib.request.ProxyHandler({'http': 'http://'})
    opener = urllib.request.build_opener(
        urllib.request.HTTPCookieProcessor(cj))
    opener.addheaders = [('User-Agent',
                          'Mozilla/5.0 (iPad; U; CPU OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5'),
                         ('Referer',
                          refer),
                         ('Host', refer),
                         ('Cookie', cookie)]
    urllib.request.install_opener(opener)

    html_bytes = urllib.request.urlopen(url).read()
    cj.save(ignore_discard=True, ignore_expires=True)
    return html_bytes


def makeUrl(product_id, page_index):
    url_head = 'https://rate.tmall.com/list_detail_rate.htm?itemId=' + \
        str(product_id)
    url_middle1 = '&'.join(['spuId=0', 'sellerId=2471189597',
                           'order=1', 'currentPage=' + str(page_index)])
    url_middle2 = '&'.join(['append=0', 'content=0',
                           'tagId=', 'posi=', 'picture='])
    url_end = '&'.join(['ua=162UW5TcyMNYQwiAiwTR3tCf0J/QnhEcUpkMmQ=|Um5Ockt/R3pEf0R4THJGeS8=|U2xMHDJ+H2QJZwBxX39RaVd5WXcxUDZKOxVDFQ==|VGhXd1llXGhQbVNoU29bZVFuWWRGfEV8SHRNdU5wTXFNcU97VQM=|VWldfS0RMQ02AzwcIBk5FydcNRQwFClYZgs+S3ZKZDJk|VmhIGCQbIQE6AiIeIhYrCzQBOgQkGCwTLg4yDzoHJxsuFi4OMg83DFoM|V25OHjAePgQ5BycbIRshATkMMmQy|WGBAED4QMGBbYVh4RH5LcScHOho0GjoONwJUAg==|WWBAED4QMAwxDDUVKRMmGjoGPwo+aD4=|WmJCEjwSMmJWaFJyTnNLfykJNBQ6FDQLNg04bjg=|W2JCEjwSMgY/AiIeIxkgAD8BPQZQBg==|XGREFDoUNGRQblR0SHRLdSMDPh4wHj4BOw86bDo=|XWRZZER5WWZGekN/X2FZY0N6WmZbe0dyUmdHe0BgW3tHfig=',
                       'isg=As_PEl7aytrfsM-V11BS-y6NX2pBRCMWUfudD-Hd6T5LsO6y6MWrZ4HixFvw',
                       '_ksTS=1495036622326_858',
                       'callback=jsonp420'])
    return '&'.join([url_head, url_middle1, url_middle2, url_end])

def getRates(product_id, title):
    orates = []
    if os.path.exists('天貓.xlsx'):
        wb = openpyxl.load_workbook('天貓.xlsx')
    else:
        wb = openpyxl.Workbook()
    try:
        sheet = wb.create_sheet(title=title)
    except Exception:
        print(title)
        sheet = wb.create_sheet(title=str(product_id))

    datafile = shelve.open('data')
    escapedCount = 0
    count = 0
    for i in range(1, 999):
        # time.sleep(2)
        url = makeUrl(product_id, i)
        # print(url)
        data = getHtml(url, 'rate.tmall.com')
        soup = BeautifulSoup(data)
        jsonp = soup.select_one('p')
        try:
            jsonp420 = jsonp.getText()
        except Exception:
            print('escaping', str(i))
            escapedCount += 1
            continue

        Regex = re.compile(r'jsonp(.*?)\(\{(.*?)\}\)')
    # print(jsonp420.startswith('jsonp420('))
        reg = Regex.search(jsonp420)
        if reg is None:
            print('escaping', str(i))
            escapedCount += 1
            continue
        json_data = '{' + reg.group(2) + '}'
        datas = json.loads(json_data)
        if('rateDetail' not in datas.keys()):
            print('escaping', str(i))
            escapedCount += 1
            continue
        rates = datas['rateDetail']['rateList']
        simply_rates = []
        if(rates == orates):
            break
        else:
            orates = rates
        index = 0
        for rate in rates:
            content, date, reply = (rate.get('rateContent', None), rate.get(
                'rateDate', None), rate.get('reply', None))
            print(content, ',', reply, ',', date)
            simply_rates.append((content, reply, date))
            sheet['A' + str(index + count + 1)] = str(date)
            sheet['B' + str(index + count + 1)] = str(content)
            sheet['C' + str(index + count + 1)] = str(reply)
            index += 1
        count += len(rates)
        osimply_rates = datafile.get('rate_' + title, [])
        datafile['rate_' + title] = osimply_rates + simply_rates
    wb.save('天貓.xlsx')
    datafile['rate_escapedCount_' + title] = escapedCount
    datafile.close()
    print('Everying done.')
    print('Escaped', escapedCount, 'times.')


def getIds(filename):
    datafile = shelve.open('data')
    htmldata = open('index/' + filename)
    soup = BeautifulSoup(htmldata)
    products = soup.select('li[class="product"]')
    product_ids = []
    for product in products:
        product_id = product.get('data-itemid')
        product_title = product.select_one('h3 a').getText()
        print(product_id, product_title)
        product_ids.append((product_id, product_title))
    datafile['product_ids_' + filename] = product_ids
    datafile.close()


getIds('1.htm')
datafile = shelve.open('data')
for pid, title in datafile['product_ids_1.htm']:
    # print(pid, title.split()[0])
    getRates(pid, title.split()[0])
datafile.close()

getIds('2.htm')
datafile = shelve.open('data')
for pid, title in datafile['product_ids_2.htm']:
    # print(pid, title.split()[0])
    getRates(pid, title.split()[0])
datafile.close()

getIds('3.htm')
datafile = shelve.open('data')
for pid, title in datafile['product_ids_3.htm']:
    # print(pid, title.split()[0])
    getRates(pid, title.split()[0])
datafile.close()
