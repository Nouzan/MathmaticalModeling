import openpyxl
import shelve


def weeklydata():
    file = shelve.open('data')
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title = '2017'
    tabcount = 0
    for table in file['2017']:
        count = 0
        for row in table:
            count = count + 1
            for i in range(len(row)):
                sheet.cell(row=count + tabcount, column=i + 1, value=row[i])
                tabcount = tabcount + len(table)
    wb.save('1.xlsx')
    file.close()


def type_season(keyword):
    datafile = shelve.open('data')
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title = keyword
    secount = 0
    for season, datas in datafile[keyword].items():
        sheet['A' + str(secount + 1)] = season
        sheet.merge_cells('A' + str(secount + 1) + ':A' +
                          str(secount + len(datas)))
        i = 0
        for k, (v1, v2) in datas.items():
            sheet['B' + str(i + secount + 1)] = k
            sheet['C' + str(i + secount + 1)] = v1
            sheet['D' + str(i + secount + 1)] = v2
            i += 1
        secount += len(datas)
    wb.save(keyword + '.xlsx')
    datafile.close()


type_season('chantype_season')
