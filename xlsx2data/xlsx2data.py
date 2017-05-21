import openpyxl
import pprint

data = {
    'key': 'data1',  # primary key
    'item2': 'data2',
    'item3': 'data3'
}
sheet = {
    'title': 'sheet_title',
    'datas': [{
        'key': 'data1',  # primary key
        'item2': 'data2',
        'item3': 'data3'
    }, {
        'key': 'data1',  # primary key
        'item2': 'data2',
        'item3': 'data3'
    }, {
        'key': 'data1',  # primary key
        'item2': 'data2',
        'item3': 'data3'}
    ]
}


def xlsx2data_onesheet(filename,
                       key_row=1,
                       key_column=1,
                       rown=2, column=1,
                       sheet_index=0):
    wb = openpyxl.load_workbook(filename, data_only=True)
    sheet_names = wb.get_sheet_names()
    raw_sheet = wb.get_sheet_by_name(sheet_names[sheet_index])
    sheet = {}
    sheet['title'] = sheet_names[sheet_index]
    datas = []
    keys = []
    primary_key = raw_sheet.cell(row=key_row, column=key_column).value
    merged_cells_groups = []
    for cells_range in raw_sheet.merged_cell_ranges:
        merged_cells = []
        for cell in raw_sheet[cells_range]:
            merged_cells.append(cell[0])
        merged_cells_groups.append(merged_cells)
    rowcount = 0
    for row in raw_sheet.rows:
        if rowcount < rown-1:
            rowcount += 1
            continue
        data = {}
        subdatas = {}
        for cell in row:
            if cell.col_idx != key_column:
                itemtitle = raw_sheet.cell(row=key_row, column=cell.col_idx).value
                if itemtitle is not None:
                    value = cell.value
                    for cells in merged_cells_groups:
                        if cell in cells:
                            value = cells[0].value
                    subdatas[itemtitle] = value
        keyvalue = row[key_column-1].value
        if keyvalue not in keys:
            keys.append(keyvalue)
        data['datas'] = subdatas
        data['key'] = keyvalue
        datas.append(data)
        # print(datas)
    sheet['items'] = datas
    sheet['key_title'] = primary_key
    sheet['key_list'] = keys
    pprint.pprint(sheet)
    return sheet
